# bill.py
# -*- coding: utf-8 -*-

import re
import io
import pandas as pd
import streamlit as st

st.set_page_config(page_title="Sales Bill Converter", layout="wide")

DATE_RE = re.compile(r"(\d{1,2}/\d{1,2}/\d{4})")


# ----------------------------
# Column letter helpers
# ----------------------------
def idx_to_col(n: int) -> str:
    n += 1
    s = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(ord("A") + r) + s
    return s


def col_to_idx(col: str) -> int:
    col = col.strip().upper()
    n = 0
    for ch in col:
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n - 1


# ----------------------------
# Helpers
# ----------------------------
def as_str(x):
    try:
        if pd.isna(x):
            return ""
    except Exception:
        pass
    if x is None:
        return ""
    return str(x).strip()


def to_float(v):
    try:
        if pd.isna(v):
            return None
    except Exception:
        pass
    if isinstance(v, (int, float)):
        if isinstance(v, float) and (v != v):
            return None
        return float(v)
    s = as_str(v)
    if s == "":
        return None
    s = s.replace(",", "")
    try:
        f = float(s)
        if f != f:
            return None
        return f
    except Exception:
        return None


def normalize_time(v):
    if v is None or v == "":
        return ""
    s = as_str(v)
    m = re.match(r"^(\d{1,2}:\d{2})", s)
    return m.group(1) if m else s


def looks_like_total_text(item_text: str) -> bool:
    s = as_str(item_text).upper()
    if s == "":
        return False
    t = as_str(item_text)
    return ("TOTAL" in s) or ("‡∏£‡∏ß‡∏°" in t) or ("‡∏¢‡∏≠‡∏î‡∏£‡∏ß‡∏°" in t)


def normalize_payment(text: str) -> str:
    s = as_str(text)
    if s == "":
        return ""
    s2 = s.replace(" ", "")
    if "‡πÄ‡∏á‡∏¥‡∏ô‡∏™‡∏î" in s2:
        return "‡πÄ‡∏á‡∏¥‡∏ô‡∏™‡∏î"
    if "‡∏™‡πÅ‡∏Å‡∏ô‡∏à‡πà‡∏≤‡∏¢" in s2 or "‡∏™‡πÅ‡∏Å‡∏ô" in s2:
        return "‡∏™‡πÅ‡∏Å‡∏ô‡∏à‡πà‡∏≤‡∏¢"
    if "‡∏Ñ‡∏ô‡∏•‡∏∞‡∏Ñ‡∏£‡∏∂‡πà‡∏á" in s2:
        return "‡∏Ñ‡∏ô‡∏•‡∏∞‡∏Ñ‡∏£‡∏∂‡πà‡∏á"
    return ""


def df_to_excel_bytes(df: pd.DataFrame):
    bio = io.BytesIO()
    with pd.ExcelWriter(bio, engine="xlsxwriter") as writer:
        df.to_excel(writer, index=False, sheet_name="sales")
    return bio.getvalue()


def is_numeric_like(s: str) -> bool:
    return bool(re.fullmatch(r"\d+", as_str(s)))


def is_barcode_like(s: str) -> bool:
    # barcode ‡∏°‡∏±‡∏Å‡∏¢‡∏≤‡∏ß 10-14+ ‡∏´‡∏•‡∏±‡∏Å (‡πÑ‡∏ó‡∏¢‡∏û‡∏ö‡∏ö‡πà‡∏≠‡∏¢ 13 ‡∏´‡∏•‡∏±‡∏Å‡πÄ‡∏£‡∏¥‡πà‡∏° 885)
    s = as_str(s)
    return bool(re.fullmatch(r"\d{10,}", s))


def is_bill_no_text(s: str) -> bool:
    """
    ‚úÖ ‡∏ï‡∏≤‡∏°‡∏ó‡∏µ‡πà‡∏Ñ‡∏∏‡∏ì‡∏¢‡πâ‡∏≥: ‡πÄ‡∏•‡∏Ç‡∏ö‡∏¥‡∏• = ‡∏ï‡∏±‡∏ß‡πÄ‡∏•‡∏Ç 6 ‡∏´‡∏•‡∏±‡∏Å‡πÄ‡∏ó‡πà‡∏≤‡∏ô‡∏±‡πâ‡∏ô
    (‡πÅ‡∏•‡∏∞‡∏Å‡∏±‡∏ô barcode 10+ ‡∏´‡∏•‡∏±‡∏Å‡πÑ‡∏ß‡πâ‡∏î‡πâ‡∏ß‡∏¢‡πÄ‡∏ú‡∏∑‡πà‡∏≠‡∏Ñ‡∏≠‡∏•‡∏±‡∏°‡∏ô‡πå‡πÄ‡∏î‡∏µ‡∏¢‡∏ß‡∏Å‡∏±‡∏ô‡∏°‡∏µ‡∏ö‡∏≤‡∏£‡πå‡πÇ‡∏Ñ‡πâ‡∏î)
    """
    s = as_str(s)
    if not re.fullmatch(r"\d{6}", s):
        return False
    if is_barcode_like(s):
        return False
    return True


def make_unique_bill_id(machine_name: str, bill_no: str) -> str:
    m = as_str(machine_name)
    b = as_str(bill_no)
    return f"{m}-{b}" if m else f"-{b}"


# ----------------------------
# Header extraction (AUTO)
# ----------------------------
def extract_header_info_auto(rows, bill_col_idx=0, scan_cols=200, max_header_rows=120):
    first_bill_row = None
    for r in range(min(max_header_rows, len(rows))):
        row = rows[r] or []
        if bill_col_idx < len(row) and is_bill_no_text(row[bill_col_idx]):
            first_bill_row = r
            break

    header_end = first_bill_row if first_bill_row is not None else min(max_header_rows, len(rows))

    header_cells = []
    for r in range(header_end):
        row = rows[r] or []
        for c in range(min(scan_cols, len(row))):
            s = as_str(row[c])
            if s:
                header_cells.append(s)

    all_dates = []
    for s in header_cells:
        all_dates.extend(DATE_RE.findall(s))
    date_from = all_dates[0] if len(all_dates) >= 1 else ""
    date_to = all_dates[1] if len(all_dates) >= 2 else (date_from if date_from else "")

    candidates = [s for s in header_cells if re.fullmatch(r"[A-Za-z]{4,30}", s)]
    machine_name = max(candidates, key=len) if candidates else ""

    return machine_name, date_from, date_to


# ----------------------------
# Core parser
# ----------------------------
def parse_rows_to_sales(rows, colmap, header_info, next_item_idx=None, stop_on_empty_rows=10):
    machine_name, date_from, date_to = header_info

    out = []
    empty_run = 0

    current_bill = ""
    current_time = ""
    current_payment = ""
    payment_by_bill = {}

    for _, row in enumerate(rows, start=1):
        if row is None:
            row = []

        if all(as_str(c) == "" for c in row):
            empty_run += 1
            if empty_run >= stop_on_empty_rows:
                break
            continue
        empty_run = 0

        # ‡∏≠‡πà‡∏≤‡∏ô‡πÄ‡∏ß‡∏•‡∏≤/‡∏ß‡∏¥‡∏ò‡∏µ‡∏à‡πà‡∏≤‡∏¢‡∏Ç‡∏≠‡∏á‡πÅ‡∏ñ‡∏ß‡∏ô‡∏µ‡πâ‡∏Å‡πà‡∏≠‡∏ô (‡∏¢‡∏∑‡∏ô‡∏¢‡∏±‡∏ô‡∏´‡∏±‡∏ß‡∏ö‡∏¥‡∏•)
        raw_time_now = normalize_time(row[colmap["time"]]) if colmap["time"] < len(row) else ""
        raw_pay_now = as_str(row[colmap["pay"]]) if colmap["pay"] < len(row) else ""
        pay_now = normalize_payment(raw_pay_now)

        # ‚úÖ ‡∏≠‡∏±‡∏õ‡πÄ‡∏î‡∏ï‡πÄ‡∏•‡∏Ç‡∏ö‡∏¥‡∏• ‡πÄ‡∏â‡∏û‡∏≤‡∏∞‡∏´‡∏±‡∏ß‡∏ö‡∏¥‡∏•‡∏à‡∏£‡∏¥‡∏á + ‡∏ï‡πâ‡∏≠‡∏á‡πÄ‡∏õ‡πá‡∏ô‡πÄ‡∏•‡∏Ç 6 ‡∏´‡∏•‡∏±‡∏Å‡πÄ‡∏ó‡πà‡∏≤‡∏ô‡∏±‡πâ‡∏ô
        raw_bill = as_str(row[colmap["bill_no"]]) if colmap["bill_no"] < len(row) else ""
        if raw_bill != "" and is_bill_no_text(raw_bill) and (raw_time_now != "" or pay_now != ""):
            current_bill = raw_bill

        if current_bill == "":
            continue

        # ‡πÉ‡∏ä‡πâ‡∏Ñ‡πà‡∏≤‡πÄ‡∏ß‡∏•‡∏≤/‡∏ß‡∏¥‡∏ò‡∏µ‡∏à‡πà‡∏≤‡∏¢
        if raw_time_now != "":
            current_time = raw_time_now

        if pay_now != "":
            current_payment = pay_now
            payment_by_bill[current_bill] = pay_now
        else:
            if current_bill in payment_by_bill:
                current_payment = payment_by_bill[current_bill]

        # item
        item = as_str(row[colmap["item"]]) if colmap["item"] < len(row) else ""
        item2 = ""
        if next_item_idx is not None and next_item_idx < len(row):
            item2 = as_str(row[next_item_idx])

        # ‡∏Ç‡πâ‡∏≤‡∏° TOTAL ‡πÄ‡∏î‡∏¥‡∏°‡πÉ‡∏ô‡πÑ‡∏ü‡∏•‡πå
        if looks_like_total_text(item) or looks_like_total_text(item2):
            continue

        # ‡∏ñ‡πâ‡∏≤ item ‡∏ß‡πà‡∏≤‡∏á ‡πÉ‡∏ä‡πâ item2
        if item == "" and item2 != "":
            item = item2

        # ‡∏Å‡∏±‡∏ô‡πÄ‡∏Ñ‡∏™ item ‡πÄ‡∏ú‡∏•‡∏≠‡πÄ‡∏õ‡πá‡∏ô‡πÄ‡∏•‡∏Ç‡∏ö‡∏¥‡∏• (6 ‡∏´‡∏•‡∏±‡∏Å) -> ‡∏ñ‡∏∑‡∏≠‡∏ß‡πà‡∏≤‡πÑ‡∏°‡πà‡πÉ‡∏ä‡πà‡∏ä‡∏∑‡πà‡∏≠‡∏™‡∏¥‡∏ô‡∏Ñ‡πâ‡∏≤
        if item != "" and is_bill_no_text(item):
            if item2 != "" and (not is_bill_no_text(item2)):
                item = item2
            else:
                continue

        if item == "":
            continue

        qty = to_float(row[colmap["qty"]]) if colmap["qty"] < len(row) else None
        price = to_float(row[colmap["price"]]) if colmap["price"] < len(row) else None
        amount = to_float(row[colmap["amount"]]) if colmap["amount"] < len(row) else None

        if qty is None and price is None and amount is None:
            continue

        # discount: ‡πÄ‡∏â‡∏û‡∏≤‡∏∞‡∏ö‡∏£‡∏£‡∏ó‡∏±‡∏î‡∏™‡∏¥‡∏ô‡∏Ñ‡πâ‡∏≤ ‡∏ñ‡πâ‡∏≤‡∏¢‡∏≠‡∏î‡∏£‡∏ß‡∏°‡∏™‡∏¥‡∏ô‡∏Ñ‡πâ‡∏≤ (line_amount) ‡∏ï‡∏¥‡∏î‡∏•‡∏ö
        discount = 0.0
        if amount is not None and amount < 0:
            discount = float(abs(amount))

        out.append(
            {
                "machine_name": machine_name,
                "unique_bill_id": make_unique_bill_id(machine_name, current_bill),
                "date_from": date_from,
                "date_to": date_to,
                "bill_no": current_bill,
                "time": current_time,
                "payment_method": current_payment,
                "item": item,
                "qty": qty,
                "price": price,
                "discount": discount,
                "line_amount": amount,
                "bill_total": None,
            }
        )

    df = pd.DataFrame(out)
    if df.empty:
        return df

    df["_row_order"] = df.groupby("bill_no").cumcount()

    bill_sum = (
        df.groupby("bill_no")["line_amount"]
        .sum(min_count=1)
        .rename("bill_total")
        .reset_index()
    )

    last_time = (
        df.groupby("bill_no")["time"]
        .agg(lambda s: next((x for x in s[::-1] if as_str(x) != ""), ""))
        .rename("time")
        .reset_index()
    )

    last_pay = (
        df.groupby("bill_no")["payment_method"]
        .agg(lambda s: next((x for x in s[::-1] if as_str(x) != ""), ""))
        .rename("payment_method")
        .reset_index()
    )

    # TOTAL row ‡∏ï‡πà‡∏≠‡∏ö‡∏¥‡∏•
    total_rows = bill_sum.merge(last_time, on="bill_no", how="left").merge(last_pay, on="bill_no", how="left")
    total_rows["item"] = "TOTAL"
    total_rows["qty"] = None
    total_rows["price"] = None
    total_rows["line_amount"] = None
    total_rows["discount"] = None  # TOTAL ‡πÑ‡∏°‡πà‡πÉ‡∏™‡πà discount
    total_rows["_row_order"] = 10**9
    total_rows["machine_name"] = machine_name
    total_rows["unique_bill_id"] = total_rows["bill_no"].apply(lambda b: make_unique_bill_id(machine_name, b))
    total_rows["date_from"] = date_from
    total_rows["date_to"] = date_to

    df_out = pd.concat([df, total_rows], ignore_index=True, sort=False)
    df_out["__bill_sort__"] = pd.to_numeric(df_out["bill_no"], errors="coerce")

    df_out = (
        df_out.sort_values(by=["__bill_sort__", "_row_order"], ascending=[True, True])
        .drop(columns=["__bill_sort__", "_row_order"])
    )

    return df_out[
        [
            "machine_name",
            "unique_bill_id",
            "date_from",
            "date_to",
            "bill_no",
            "time",
            "payment_method",
            "item",
            "qty",
            "price",
            "discount",
            "line_amount",
            "bill_total",
        ]
    ]


# ----------------------------
# File reading
# ----------------------------
def read_rows_from_upload(uploaded_file, sheet_name=None):
    name = uploaded_file.name.lower()
    if name.endswith(".xlsx"):
        from openpyxl import load_workbook

        wb = load_workbook(uploaded_file, data_only=True)
        if sheet_name is None:
            sheet_name = wb.sheetnames[0]
        ws = wb[sheet_name]
        max_col = ws.max_column or 1
        rows = [list(r) for r in ws.iter_rows(min_row=1, max_col=max_col, values_only=True)]
        return rows, wb.sheetnames, sheet_name
    elif name.endswith(".xls"):
        xls = pd.ExcelFile(uploaded_file, engine="xlrd")
        if sheet_name is None:
            sheet_name = xls.sheet_names[0]
        df_raw = pd.read_excel(xls, sheet_name=sheet_name, header=None, engine="xlrd")
        rows = df_raw.values.tolist()
        return rows, xls.sheet_names, sheet_name
    else:
        raise ValueError("‡∏£‡∏≠‡∏á‡∏£‡∏±‡∏ö‡πÄ‡∏â‡∏û‡∏≤‡∏∞ .xlsx / .xls")


# ----------------------------
# UI
# ----------------------------
st.title("üßæ ‡πÅ‡∏õ‡∏•‡∏á‡πÑ‡∏ü‡∏•‡πå‡∏ö‡∏¥‡∏• (TOTAL ‡πÉ‡∏´‡∏°‡πà + ‡∏ß‡∏¥‡∏ò‡∏µ‡∏à‡πà‡∏≤‡∏¢ + ‡∏ä‡∏∑‡πà‡∏≠‡πÄ‡∏Ñ‡∏£‡∏∑‡πà‡∏≠‡∏á/‡∏ß‡∏±‡∏ô‡∏ó‡∏µ‡πà‡∏à‡∏≤‡∏Å‡∏´‡∏±‡∏ß‡πÑ‡∏ü‡∏•‡πå + Unique Bill ID + Discount)")

uploaded_files = st.file_uploader(
    "‡∏≠‡∏±‡∏õ‡πÇ‡∏´‡∏•‡∏î‡πÑ‡∏ü‡∏•‡πå Excel (.xlsx / .xls) ‡πÑ‡∏î‡πâ‡∏´‡∏•‡∏≤‡∏¢‡πÑ‡∏ü‡∏•‡πå",
    type=["xlsx", "xls"],
    accept_multiple_files=True,
)

if not uploaded_files:
    st.info("‡∏Å‡∏£‡∏∏‡∏ì‡∏≤‡∏≠‡∏±‡∏õ‡πÇ‡∏´‡∏•‡∏î‡πÑ‡∏ü‡∏•‡πå‡∏Å‡πà‡∏≠‡∏ô‡∏Ñ‡∏£‡∏±‡∏ö")
    st.stop()

rows0, _, _ = read_rows_from_upload(uploaded_files[0], sheet_name=None)
max_cols_detected = max((len(r) for r in rows0), default=1)
col_letters = [idx_to_col(i) for i in range(max_cols_detected)]

st.subheader("‡∏ï‡∏±‡πâ‡∏á‡∏Ñ‡πà‡∏≤‡∏Ñ‡∏≠‡∏•‡∏±‡∏°‡∏ô‡πå‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏•‡∏ö‡∏¥‡∏• (‡πÉ‡∏ä‡πâ‡∏£‡πà‡∏ß‡∏°‡∏Å‡∏±‡∏ô‡∏ó‡∏∏‡∏Å‡πÑ‡∏ü‡∏•‡πå)")


def safe_index(colname: str) -> int:
    i = col_to_idx(colname)
    return i if 0 <= i < len(col_letters) else 0


c0, c1, c2, c3, c4, c5, c6 = st.columns(7)
with c0:
    bill_col = st.selectbox("‡πÄ‡∏•‡∏Ç‡∏ö‡∏¥‡∏• (bill_no)", col_letters, index=safe_index("A"))
with c1:
    time_col = st.selectbox("‡πÄ‡∏ß‡∏•‡∏≤ (time)", col_letters, index=safe_index("L"))
with c2:
    pay_col = st.selectbox("‡∏ß‡∏¥‡∏ò‡∏µ‡∏à‡πà‡∏≤‡∏¢", col_letters, index=safe_index("D"))
with c3:
    item_col = st.selectbox("‡∏™‡∏¥‡∏ô‡∏Ñ‡πâ‡∏≤ (item)", col_letters, index=safe_index("A"))
with c4:
    qty_col = st.selectbox("‡∏à‡∏≥‡∏ô‡∏ß‡∏ô (qty)", col_letters, index=safe_index("G"))
with c5:
    price_col = st.selectbox("‡∏£‡∏≤‡∏Ñ‡∏≤ (price)", col_letters, index=safe_index("I"))
with c6:
    amt_col = st.selectbox("‡∏¢‡∏≠‡∏î‡∏ö‡∏£‡∏£‡∏ó‡∏±‡∏î (line_amount)", col_letters, index=safe_index("K"))

st.caption("‡∏ñ‡πâ‡∏≤ '‡∏™‡∏¥‡∏ô‡∏Ñ‡πâ‡∏≤' ‡∏ö‡∏≤‡∏á‡πÅ‡∏ñ‡∏ß‡∏´‡∏•‡∏∏‡∏î‡πÑ‡∏õ‡πÄ‡∏õ‡πá‡∏ô‡πÄ‡∏•‡∏Ç‡∏ö‡∏¥‡∏• ‡πÉ‡∏´‡πâ‡πÄ‡∏•‡∏∑‡∏≠‡∏Å‡∏Ñ‡∏≠‡∏•‡∏±‡∏°‡∏ô‡πå‡∏™‡∏≥‡∏£‡∏≠‡∏á‡∏ó‡∏µ‡πà‡πÄ‡∏õ‡πá‡∏ô‡∏ä‡∏∑‡πà‡∏≠‡∏™‡∏¥‡∏ô‡∏Ñ‡πâ‡∏≤")
next_item_col = st.selectbox("‡∏Ñ‡∏≠‡∏•‡∏±‡∏°‡∏ô‡πå‡∏ä‡∏∑‡πà‡∏≠‡∏™‡∏¥‡∏ô‡∏Ñ‡πâ‡∏≤ (‡∏™‡∏≥‡∏£‡∏≠‡∏á / ‡∏ñ‡∏±‡∏î‡πÑ‡∏õ)", ["(‡πÑ‡∏°‡πà‡πÉ‡∏ä‡πâ)"] + col_letters, index=0)
next_item_idx = None if next_item_col == "(‡πÑ‡∏°‡πà‡πÉ‡∏ä‡πâ)" else col_to_idx(next_item_col)

colmap = {
    "bill_no": col_to_idx(bill_col),
    "time": col_to_idx(time_col),
    "pay": col_to_idx(pay_col),
    "item": col_to_idx(item_col),
    "qty": col_to_idx(qty_col),
    "price": col_to_idx(price_col),
    "amount": col_to_idx(amt_col),
}

st.subheader("‡πÄ‡∏•‡∏∑‡∏≠‡∏Å‡∏ä‡∏µ‡∏ó‡∏ï‡πà‡∏≠‡πÑ‡∏ü‡∏•‡πå (‡∏ñ‡πâ‡∏≤‡∏ï‡πâ‡∏≠‡∏á‡∏Å‡∏≤‡∏£)")
sheet_choice = {}
for uf in uploaded_files:
    _, sheetnames, default_sheet = read_rows_from_upload(uf, sheet_name=None)
    with st.expander(f"‡πÑ‡∏ü‡∏•‡πå: {uf.name}  (‡∏Ñ‡πà‡∏≤‡πÄ‡∏£‡∏¥‡πà‡∏°‡∏ï‡πâ‡∏ô: {default_sheet})", expanded=False):
        sheet_choice[uf.name] = st.selectbox(
            f"‡πÄ‡∏•‡∏∑‡∏≠‡∏Å‡∏ä‡∏µ‡∏ó‡∏™‡∏≥‡∏´‡∏£‡∏±‡∏ö {uf.name}",
            sheetnames,
            index=sheetnames.index(default_sheet),
            key=f"sheet_{uf.name}",
        )

dfs = []
header_preview = []

for uf in uploaded_files:
    rows, _, used_sheet = read_rows_from_upload(uf, sheet_name=sheet_choice.get(uf.name))
    header_info = extract_header_info_auto(rows, bill_col_idx=colmap["bill_no"], scan_cols=200)
    machine_name, date_from, date_to = header_info

    df = parse_rows_to_sales(rows, colmap, header_info=header_info, next_item_idx=next_item_idx)
    if df.empty:
        continue

    dfs.append(df)
    header_preview.append(
        f"- {uf.name} ‚Üí machine: {machine_name or '(‡πÑ‡∏°‡πà‡∏û‡∏ö)'} | {date_from or '-'} ‡∏ñ‡∏∂‡∏á {date_to or '-'} | sheet: {used_sheet}"
    )

st.subheader("‡∏´‡∏±‡∏ß‡πÑ‡∏ü‡∏•‡πå‡∏ó‡∏µ‡πà‡∏≠‡πà‡∏≤‡∏ô‡πÑ‡∏î‡πâ (‡∏ï‡πà‡∏≠‡πÑ‡∏ü‡∏•‡πå)")
st.write("\n".join(header_preview) if header_preview else "‡πÑ‡∏°‡πà‡∏û‡∏ö‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏•‡∏´‡∏±‡∏ß‡πÑ‡∏ü‡∏•‡πå")

if not dfs:
    st.error("‡πÑ‡∏°‡πà‡πÄ‡∏à‡∏≠‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏•‡∏Ç‡∏≤‡∏¢‡πÉ‡∏ô‡∏ó‡∏∏‡∏Å‡πÑ‡∏ü‡∏•‡πå ‚Äî ‡πÄ‡∏ä‡πá‡∏Ñ‡∏Ñ‡∏≠‡∏•‡∏±‡∏°‡∏ô‡πå‡∏ó‡∏µ‡πà‡πÄ‡∏•‡∏∑‡∏≠‡∏Å")
    st.stop()

df_all = pd.concat(dfs, ignore_index=True)

# ‡πÄ‡∏õ‡∏•‡∏µ‡πà‡∏¢‡∏ô‡∏ä‡∏∑‡πà‡∏≠‡∏Ñ‡∏≠‡∏•‡∏±‡∏°‡∏ô‡πå‡∏ï‡∏≤‡∏°‡∏ó‡∏µ‡πà‡∏Ç‡∏≠
df_all = df_all.rename(
    columns={
        "line_amount": "‡∏¢‡∏≠‡∏î‡∏£‡∏ß‡∏°‡∏™‡∏¥‡∏ô‡∏Ñ‡πâ‡∏≤",
        "bill_total": "‡∏¢‡∏≠‡∏î‡∏£‡∏ß‡∏°‡∏ö‡∏¥‡∏•",
    }
)

st.subheader("‡∏ú‡∏•‡∏•‡∏±‡∏û‡∏ò‡πå‡∏£‡∏ß‡∏° (‡∏ó‡∏∏‡∏Å‡πÑ‡∏ü‡∏•‡πå)")
st.write(f"‡∏û‡∏ö‡∏£‡∏≤‡∏¢‡∏Å‡∏≤‡∏£‡∏£‡∏ß‡∏°: **{len(df_all):,}** ‡πÅ‡∏ñ‡∏ß | ‡∏û‡∏ö‡πÄ‡∏•‡∏Ç‡∏ö‡∏¥‡∏•‡∏£‡∏ß‡∏°: **{df_all['bill_no'].nunique():,}** ‡∏ö‡∏¥‡∏•")
st.dataframe(df_all.head(300), use_container_width=True)

csv_bytes = df_all.to_csv(index=False).encode("utf-8-sig")
xlsx_bytes = df_to_excel_bytes(df_all)

cA, cB = st.columns(2)
with cA:
    st.download_button("‚¨áÔ∏è ‡∏î‡∏≤‡∏ß‡∏ô‡πå‡πÇ‡∏´‡∏•‡∏î CSV (‡∏£‡∏ß‡∏°)", data=csv_bytes, file_name="sales_clean_all.csv", mime="text/csv")
with cB:
    st.download_button(
        "‚¨áÔ∏è ‡∏î‡∏≤‡∏ß‡∏ô‡πå‡πÇ‡∏´‡∏•‡∏î Excel (‡∏£‡∏ß‡∏°)",
        data=xlsx_bytes,
        file_name="sales_clean_all.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
